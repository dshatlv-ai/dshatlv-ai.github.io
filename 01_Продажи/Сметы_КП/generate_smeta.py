#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор смет (Lokālā tāme) в формате образца Smeta_ieejas_zona_PASUTITAJS.xlsx.

Использование:
    python3 generate_smeta.py              # интерактивный режим
    python3 generate_smeta.py --catalog    # показать справочник работ
    python3 generate_smeta.py --demo       # собрать тестовую смету для проверки
    python3 generate_smeta.py --out ПАПКА  # куда сохранять готовые сметы

Справочник работ лежит рядом со скриптом: rabota_catalog.xlsx.
Его можно пополнять прямо в Excel (добавлять строки) или из интерактивного
режима — при вводе позиции, которой нет в справочнике, скрипт предложит
сохранить её туда.
"""

import argparse
import datetime as _dt
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Нужен пакет openpyxl:  pip install openpyxl")

# ---------------------------------------------------------------------------
# РЕКВИЗИТЫ И ТЕКСТЫ ШАПКИ — правь под себя здесь
# ---------------------------------------------------------------------------

TITLE = "Lokālā tāme"
EXECUTOR = "Izpildītājs: Daniils Šatalovs (saimnieciskās darbības veicējs)"
DEFAULT_CUSTOMER = 'Pasūtītājs: DzīB "Vārnu iela 8, 8A, 8B, 8C", reģ. Nr. 40008318408'

# Стандартные примечания внизу сметы (можно менять/дополнять)
DEFAULT_NOTES = [
    "• Cenas norādītas EUR, PVN netiek piemērots.",
    "• SLĒPTIE DARBI: papildu defektus pēc virsmas atvēršanas apmaksā atsevišķi pēc faktiskā darbu apjoma un materiāliem.",
    "• Apmaksa: avanss 40% pirms darbu sākuma, atlikums pēc pieņemšanas akta parakstīšanas.",
    "• Piedāvājuma derīguma termiņš — 30 dienas no sastādīšanas datuma.",
]

SIGNATURE_LEFT = "Sastādīja: ________________ Daniils Šatalovs"
SIGNATURE_RIGHT = "Pieņēma: ________________"

CATALOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rabota_catalog.xlsx")

# ---------------------------------------------------------------------------
# СТАРТОВЫЙ СПРАВОЧНИК (создаётся при первом запуске, дальше правь в Excel)
#   (Nosaukums LV, Название RU — для поиска, Mērvienība, Cena EUR, Piezīme)
# ---------------------------------------------------------------------------

STARTER_ITEMS = [
    ("Grīdas seguma demontāža, iznešana", "демонтаж пола", "m²", 15.00, "darbs"),
    ("Grīdas klons / izlīdzinošā stiedze", "стяжка", "m²", 25.00, "darbs"),
    ("Flīžu ieklāšana uz grīdas, šuvošana", "укладка плитки на пол", "m²", 45.00, "darbs"),
    ("Kāpņu pakāpienu flīzēšana", "облицовка ступеней плиткой", "t.m.", 60.00, "darbs"),
    ("Sienu krāsošana (2 kārtās)", "покраска стен", "m²", 12.00, "darbs"),
    ("Sienu špaktelēšana", "шпаклёвка", "m²", 10.00, "darbs"),
    ("Gruntēšana", "грунтовка", "m²", 8.00, "materiāls+darbs"),
    ("Būvgružu savākšana un iznešana", "вывоз мусора", "kompl.", 180.00, "darbs"),
    ("Konteinera noma un izvešana poligonā", "контейнер", "reize", 295.00, "utilizācija"),
    ("Materiālu piegāde, transports", "доставка материалов", "kompl.", 150.00, "transports"),
]

# ---------------------------------------------------------------------------
# СТИЛИ (сняты с образца)
# ---------------------------------------------------------------------------

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEAD = PatternFill("solid", start_color="FFD9D9D9")   # шапка таблицы и итоги
FILL_SECTION = PatternFill("solid", start_color="FFF2F2F2")  # заголовки секций

F9 = Font(name="Arial", size=9)
F9B = Font(name="Arial", size=9, bold=True)
F9I = Font(name="Arial", size=9, italic=True)
F10B = Font(name="Arial", size=10, bold=True)
F11B = Font(name="Arial", size=11, bold=True)
F12B = Font(name="Arial", size=12, bold=True)

COL_WIDTHS = {"A": 5, "B": 48, "C": 12, "D": 10, "E": 13, "F": 14, "G": 24}
ROW_H = 21.75  # стандартная высота строк таблицы

TABLE_HEADERS = ["Nr.", "Darbu / materiālu nosaukums", "Mērvienība", "Daudzums",
                 "Cena par vienību, EUR", "Summa, EUR", "Piezīmes"]


# ---------------------------------------------------------------------------
# СПРАВОЧНИК
# ---------------------------------------------------------------------------

def ensure_catalog(path=CATALOG_FILE):
    """Создаёт rabota_catalog.xlsx со стартовыми позициями, если его ещё нет."""
    if os.path.exists(path):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Katalogs"
    headers = ["Nr", "Nosaukums (LV)", "Название (RU)", "Mērvienība", "Cena EUR", "Piezīme"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = F9B
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, (lv, ru, unit, price, note) in enumerate(STARTER_ITEMS, 1):
        row = [i, lv, ru, unit, price, note]
        for col, v in enumerate(row, 1):
            c = ws.cell(row=i + 1, column=col, value=v)
            c.font = F9
            c.border = BORDER
            if col == 5:
                c.number_format = "0.00"
    for col, w in zip("ABCDEF", [5, 45, 32, 12, 10, 18]):
        ws.column_dimensions[col].width = w
    wb.save(path)
    print(f"Создан справочник: {path}")


def load_catalog(path=CATALOG_FILE):
    ensure_catalog(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        items.append({
            "nr": row[0],
            "name_lv": str(row[1]).strip(),
            "name_ru": str(row[2] or "").strip(),
            "unit": str(row[3] or "gab.").strip(),
            "price": float(row[4] or 0),
            "note": str(row[5] or "").strip(),
        })
    return items


def append_to_catalog(name_lv, name_ru, unit, price, note, path=CATALOG_FILE):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    r = ws.max_row + 1
    values = [r - 1, name_lv, name_ru, unit, price, note]
    for col, v in enumerate(values, 1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = F9
        c.border = BORDER
        if col == 5:
            c.number_format = "0.00"
    wb.save(path)
    print(f"Позиция «{name_lv}» добавлена в справочник.")


def find_matches(catalog, query):
    q = query.strip().lower().replace("ё", "е")
    if q.isdigit():
        exact = [it for it in catalog if str(it["nr"]) == q]
        if exact:
            return exact
    matches = []
    for it in catalog:
        hay = (it["name_lv"] + " " + it["name_ru"]).lower().replace("ё", "е")
        if q == it["name_lv"].lower() or q == it["name_ru"].lower():
            return [it]
        if q in hay:
            matches.append(it)
    return matches


def print_catalog(catalog):
    print("--- Справочник работ (rabota_catalog.xlsx) ---")
    for it in catalog:
        ru = f"  ({it['name_ru']})" if it["name_ru"] else ""
        print(f"{it['nr']:>3}. {it['name_lv']}{ru} — {it['unit']} — {it['price']:.2f} EUR")
    print("----------------------------------------------")


# ---------------------------------------------------------------------------
# СБОРКА СМЕТЫ
# ---------------------------------------------------------------------------

def _merge_row(ws, row, first="A", last="G"):
    ws.merge_cells(f"{first}{row}:{last}{row}")


def _style_range(ws, row, cols="ABCDEFG", fill=None, border=None):
    for col in cols:
        c = ws[f"{col}{row}"]
        if fill:
            c.fill = fill
        if border:
            c.border = border


def build_smeta(out_path, object_title, address, customer, sections,
                notes=None, date_str=None):
    """sections: [(заголовок_секции, [позиции])],
    позиция: dict(name, unit, qty, price, note)"""
    notes = notes if notes is not None else DEFAULT_NOTES
    date_str = date_str or _dt.date.today().strftime("%d.%m.%Y")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tāme (pasūtītājam)"
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # --- Шапка -------------------------------------------------------------
    head_lines = [
        (TITLE, F12B, center, 19.5),
        (object_title, F10B, center, 18),
        (EXECUTOR, F9, None, 15),
        (customer, F9, None, 15),
        (f"Objekta adrese: {address}", F9, None, 15),
        (f"Tāme sastādīta: {date_str}", F9, None, 15),
    ]
    for r, (text, font, align, h) in enumerate(head_lines, 1):
        _merge_row(ws, r)
        c = ws[f"A{r}"]
        c.value = text
        c.font = font
        if align:
            c.alignment = align
        ws.row_dimensions[r].height = h

    # --- Шапка таблицы -----------------------------------------------------
    r = 7
    for col, text in zip("ABCDEFG", TABLE_HEADERS):
        c = ws[f"{col}{r}"]
        c.value = text
        c.font = F9B
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 30

    # --- Секции ------------------------------------------------------------
    r += 1
    section_total_rows = []
    for sec_idx, (sec_title, items) in enumerate(sections, 1):
        # серый заголовок секции
        _merge_row(ws, r)
        _style_range(ws, r, fill=FILL_SECTION, border=BORDER)
        c = ws[f"A{r}"]
        c.value = f"{sec_idx}. {sec_title}"
        c.font = F10B
        c.alignment = left
        ws.row_dimensions[r].height = ROW_H
        r += 1

        first_item_row = r
        for item_idx, it in enumerate(items, 1):
            _style_range(ws, r, border=BORDER)
            ws[f"A{r}"].value = f"{sec_idx}.{item_idx}"
            ws[f"A{r}"].font = F9
            ws[f"A{r}"].alignment = center
            ws[f"B{r}"].value = it["name"]
            ws[f"B{r}"].font = F9
            ws[f"B{r}"].alignment = left_wrap
            ws[f"C{r}"].value = it["unit"]
            ws[f"C{r}"].font = F9
            ws[f"C{r}"].alignment = center
            for col, key in (("D", "qty"), ("E", "price")):
                c = ws[f"{col}{r}"]
                c.value = float(it[key])
                c.font = F9
                c.alignment = right
                c.number_format = "0.00"
            c = ws[f"F{r}"]
            c.value = f"=D{r}*E{r}"
            c.font = F9B
            c.alignment = right
            c.number_format = "0.00"
            c = ws[f"G{r}"]
            c.value = it.get("note", "")
            c.font = F9I
            c.alignment = left_wrap
            ws.row_dimensions[r].height = ROW_H
            r += 1

        # итог секции
        _style_range(ws, r, fill=FILL_HEAD, border=BORDER)
        ws.merge_cells(f"A{r}:E{r}")
        c = ws[f"A{r}"]
        c.value = "Kopā sadaļā:"
        c.font = F9B
        c.alignment = right
        c = ws[f"F{r}"]
        c.value = f"=SUM(F{first_item_row}:F{r - 1})"
        c.font = F9B
        c.alignment = right
        c.number_format = "0.00"
        ws.row_dimensions[r].height = ROW_H
        section_total_rows.append(r)
        r += 2  # пустая строка между секциями

    # --- KOPĀ ----------------------------------------------------------------
    _style_range(ws, r, fill=FILL_HEAD, border=BORDER)
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = "KOPĀ (PVN netiek piemērots):"
    c.font = F10B
    c.alignment = right
    c = ws[f"F{r}"]
    c.value = "=" + "+".join(f"F{tr}" for tr in section_total_rows)
    c.font = F11B
    c.alignment = right
    c.number_format = "0.00"
    ws.row_dimensions[r].height = 25.5
    r += 2

    # --- Примечания ----------------------------------------------------------
    if notes:
        c = ws[f"A{r}"]
        c.value = "Piezīmes un nosacījumi:"
        c.font = F9B
        c.alignment = left
        ws.row_dimensions[r].height = 15
        r += 1
        for note in notes:
            _merge_row(ws, r)
            c = ws[f"A{r}"]
            c.value = note
            c.font = F9
            c.alignment = left_wrap
            ws.row_dimensions[r].height = 24
            r += 1
        r += 1

    # --- Подписи ---------------------------------------------------------------
    ws.merge_cells(f"A{r}:D{r}")
    ws.merge_cells(f"E{r}:G{r}")
    ws[f"A{r}"].value = SIGNATURE_LEFT
    ws[f"A{r}"].font = F9
    ws[f"E{r}"].value = SIGNATURE_RIGHT
    ws[f"E{r}"].font = F9
    ws.row_dimensions[r].height = 15

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# ИНТЕРАКТИВНЫЙ РЕЖИМ
# ---------------------------------------------------------------------------

def _ask(prompt, default=None):
    suffix = f" [{default}]" if default else ""
    val = input(f"{prompt}{suffix}: ").strip()
    return val or (default or "")


def _parse_num(s):
    return float(s.strip().replace(",", "."))


def _slug(text):
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE).strip()
    return re.sub(r"[\s]+", "_", text)[:40] or "objekts"


def interactive(out_dir):
    catalog = load_catalog()
    print()
    print("=== Генератор смет ===")
    object_title = _ask("Название объекта (2-я строка шапки)")
    address = _ask("Адрес объекта")
    customer = _ask("Заказчик (строка целиком)", DEFAULT_CUSTOMER)
    print()
    print_catalog(catalog)
    print()
    print("Ввод позиций. Форматы строк:")
    print("  + Название секции          — начать новую секцию")
    print("  <№ или название>; <кол-во>              — позиция из справочника")
    print("  <№ или название>; <кол-во>; <цена>      — с ценой вместо каталожной")
    print("  пустая строка               — закончить ввод")
    print()

    sections = []
    current_title, current_items = None, []

    def flush():
        nonlocal current_title, current_items
        if current_items:
            sections.append((current_title or "REMONTDARBI", current_items))
        current_title, current_items = None, []

    while True:
        try:
            line = input("> ").strip()
        except EOFError:
            break
        if not line:
            break
        if line.startswith("+"):
            flush()
            current_title = line[1:].strip()
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 2:
            print("  Нужно минимум: название; количество")
            continue
        query = parts[0]
        try:
            qty = _parse_num(parts[1])
        except ValueError:
            print(f"  Не понял количество: «{parts[1]}»")
            continue
        price_override = None
        if len(parts) >= 3 and parts[2]:
            try:
                price_override = _parse_num(parts[2])
            except ValueError:
                print(f"  Не понял цену: «{parts[2]}»")
                continue

        matches = find_matches(catalog, query)
        if len(matches) > 1:
            print("  Найдено несколько, уточни:")
            for m in matches:
                print(f"    {m['nr']}. {m['name_lv']} ({m['name_ru']})")
            continue
        if matches:
            it = matches[0]
            item = {"name": it["name_lv"], "unit": it["unit"],
                    "qty": qty, "price": price_override if price_override is not None else it["price"],
                    "note": it["note"]}
        else:
            print(f"  «{query}» нет в справочнике — добавляю как новую позицию.")
            name_lv = _ask("  Название для сметы (LV)", query)
            unit = _ask("  Единица измерения (m², t.m., gab., kompl., reize)", "m²")
            if price_override is None:
                price_override = _parse_num(_ask("  Цена за единицу, EUR"))
            note = _ask("  Piezīme (darbs/materiāls/…)", "darbs")
            item = {"name": name_lv, "unit": unit, "qty": qty,
                    "price": price_override, "note": note}
            if _ask("  Сохранить в справочник? (y/n)", "y").lower().startswith(("y", "д")):
                append_to_catalog(name_lv, query, unit, price_override, note)
                catalog = load_catalog()
        current_items.append(item)
        total = item["qty"] * item["price"]
        print(f"  + {item['name']} — {item['qty']:g} {item['unit']} × {item['price']:.2f} = {total:.2f} EUR")

    flush()
    if not sections:
        sys.exit("Позиции не введены — смета не создана.")

    fname = f"Smeta_{_slug(object_title)}_{_dt.date.today().isoformat()}.xlsx"
    out_path = os.path.join(out_dir, fname)
    build_smeta(out_path, object_title, address, customer, sections)
    total = sum(i["qty"] * i["price"] for _, items in sections for i in items)
    print(f"\nГотово: {out_path}")
    print(f"KOPĀ bez PVN: {total:.2f} EUR")


# ---------------------------------------------------------------------------
# ТЕСТОВАЯ СМЕТА
# ---------------------------------------------------------------------------

def demo(out_dir):
    catalog = load_catalog()

    def cat(query, qty):
        it = find_matches(catalog, query)[0]
        return {"name": it["name_lv"], "unit": it["unit"], "qty": qty,
                "price": it["price"], "note": it["note"]}

    sections = [
        ("GRĪDA UN KĀPNES — remontdarbi", [
            cat("демонтаж пола", 11),
            cat("стяжка", 11),
            cat("облицовка ступеней плиткой", 3),
            cat("покраска стен", 20),
        ]),
        ("BŪVGRUŽI UN TRANSPORTS", [
            cat("вывоз мусора", 1),
            cat("контейнер", 1),
            cat("доставка материалов", 1),
        ]),
    ]
    out_path = os.path.join(out_dir, f"Smeta_TESTS_{_dt.date.today().isoformat()}.xlsx")
    build_smeta(
        out_path,
        object_title="Remontdarbi — testa objekts",
        address="Testa iela 1, Rīga",
        customer=DEFAULT_CUSTOMER,
        sections=sections,
    )
    total = sum(i["qty"] * i["price"] for _, items in sections for i in items)
    print(f"Тестовая смета: {out_path}")
    print(f"KOPĀ bez PVN: {total:.2f} EUR")


# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Генератор смет в формате Lokālā tāme")
    ap.add_argument("--catalog", action="store_true", help="показать справочник работ")
    ap.add_argument("--demo", action="store_true", help="собрать тестовую смету")
    ap.add_argument("--out", default=".", help="папка для готовых смет (по умолчанию текущая)")
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)
    if args.catalog:
        print_catalog(load_catalog())
    elif args.demo:
        demo(args.out)
    else:
        interactive(args.out)


if __name__ == "__main__":
    main()
